from pathlib import Path

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader


def extract_text_from_pdf(file_path: Path) -> str:
    reader = PdfReader(str(file_path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages).strip()

def extract_text_from_docx(file_path: Path) -> str:
    doc = DocxDocument(str(file_path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs).strip()

def extract_text_from_xlsx(file_path: Path) -> str:
    workbook = load_workbook(filename=str(file_path), data_only=True)
    lines = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                lines.append(" ".join(cells))
    return "\n".join(lines).strip()

EXTRACTORS = {
    "application/pdf": extract_text_from_pdf,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": extract_text_from_docx,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": extract_text_from_xlsx,
}

def extract_text(file_path: Path, content_type: str) -> str:
    extractor = EXTRACTORS.get(content_type)
    if extractor is None:
        raise ValueError(f"지원하지 않는 파일 형식입니다: {content_type}")
    return extractor(file_path)